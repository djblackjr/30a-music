"""
app/excel/exporter.py
Generate a formatted Excel workbook in exports/.
Sheets: All Events, Upcoming, Changes, By Venue
"""
import logging
from datetime import date, datetime
from pathlib import Path

logger = logging.getLogger(__name__)

EXPORTS_DIR = Path("exports")
REPORT_NAME = "30A_Live_Music_Report.xlsx"

# Colours
NAVY   = "1F3864"
TEAL   = "1B7A8A"
GOLD   = "C9A84C"
WHITE  = "FFFFFF"
LIGHT  = "F5F0E8"
GREEN_BG = "F0FDF4"
GREEN_FG = "166534"
BLUE_BG  = "E0F2FE"
BLUE_FG  = "0369A1"
RED_BG   = "FFF1F2"
RED_FG   = "9F1239"


def _header_style(ws, row: int, cols: list[str], bg: str = NAVY, fg: str = WHITE):
    from openpyxl.styles import Font, PatternFill, Alignment
    fill = PatternFill("solid", fgColor=bg)
    font = Font(bold=True, color=fg, name="Arial", size=10)
    align = Alignment(horizontal="center", vertical="center")
    for col in cols:
        cell = ws[f"{col}{row}"]
        cell.fill = fill
        cell.font = font
        cell.alignment = align


def _cell_style(cell, bold=False, color="000000", bg=None, size=9, wrap=False):
    from openpyxl.styles import Font, PatternFill, Alignment
    cell.font = Font(bold=bold, color=color, name="Arial", size=size)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(wrap_text=wrap, vertical="center")


def _write_events_sheet(ws, events: list[dict], title: str):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    ws.title = title

    # Title row
    ws.merge_cells("A1:H1")
    ws["A1"] = f"🎵 30A Music Intelligence — {title}"
    ws["A1"].font = Font(bold=True, color=WHITE, name="Arial", size=13)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Sub-title
    ws.merge_cells("A2:H2")
    ws["A2"] = f"Generated {datetime.now().strftime('%b %d, %Y %I:%M %p')}  ·  {len(events)} events"
    ws["A2"].font = Font(italic=True, color="666666", name="Arial", size=9)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    # Headers
    headers = ["Artist / Performer", "Event Name", "Date", "Day", "Time", "Venue", "Stage", "Source"]
    cols    = ["A","B","C","D","E","F","G","H"]
    for col, hdr in zip(cols, headers):
        ws[f"{col}3"] = hdr
    _header_style(ws, 3, cols, bg=TEAL)
    ws.row_dimensions[3].height = 20

    # Data rows
    today = date.today().isoformat()
    for i, ev in enumerate(events, start=4):
        ev_date = ev.get("date") or ""
        is_past = ev_date and ev_date < today
        row_bg  = "F9F9F9" if i % 2 == 0 else WHITE

        try:
            dt = datetime.strptime(ev_date, "%Y-%m-%d")
            day_str  = dt.strftime("%A")
            date_str = dt.strftime("%b %d, %Y")
        except Exception:
            day_str  = ""
            date_str = ev_date

        values = [
            ev.get("performer") or "",
            ev.get("name") or "",
            date_str,
            day_str,
            ev.get("time_start") or "",
            ev.get("venue") or "",
            ev.get("stage") or "",
            ev.get("source") or "",
        ]
        for col, val in zip(cols, values):
            cell = ws[f"{col}{i}"]
            cell.value = val
            txt_color = "AAAAAA" if is_past else "222222"
            _cell_style(cell, color=txt_color, bg=row_bg)

        ws.row_dimensions[i].height = 18

    # Column widths
    widths = [22, 30, 14, 12, 8, 20, 18, 20]
    for col, w in zip(cols, widths):
        ws.column_dimensions[col].width = w

    # Freeze panes below header
    ws.freeze_panes = "A4"


def _write_changes_sheet(ws, changes: dict):
    from openpyxl.styles import Font, PatternFill, Alignment

    ws.title = "Changes"

    ws.merge_cells("A1:E1")
    ws["A1"] = "Run-to-Run Changes"
    ws["A1"].font = Font(bold=True, color=WHITE, name="Arial", size=12)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    summary = changes.get("summary", {})
    row = 3
    for label, key, bg, fg in [
        ("🆕 New Events",      "new",     "D1FAE5", "065F46"),
        ("✏️  Changed Events", "changed", "FEF9C3", "713F12"),
        ("🗑️  Removed Events", "removed", "FEE2E2", "7F1D1D"),
        ("✅ Unchanged",       "unchanged","F1F5F9","334155"),
    ]:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = summary.get(key, 0)
        ws[f"A{row}"].font = Font(bold=True, name="Arial", size=10, color=fg)
        ws[f"A{row}"].fill = PatternFill("solid", fgColor=bg)
        ws[f"B{row}"].font = Font(bold=True, name="Arial", size=10)
        ws.row_dimensions[row].height = 20
        row += 1

    row += 1
    # New events detail
    new_events = changes.get("new", [])
    if new_events:
        ws[f"A{row}"] = "NEW EVENTS THIS RUN"
        ws[f"A{row}"].font = Font(bold=True, color=WHITE, name="Arial", size=10)
        ws[f"A{row}"].fill = PatternFill("solid", fgColor="065F46")
        row += 1
        for ev in new_events:
            ws[f"A{row}"] = ev.get("performer") or ev.get("name") or ""
            ws[f"B{row}"] = ev.get("date") or ""
            ws[f"C{row}"] = ev.get("time_start") or ""
            ws[f"D{row}"] = ev.get("venue") or ""
            ws[f"E{row}"] = ev.get("stage") or ""
            ws.row_dimensions[row].height = 16
            row += 1

    for col, w in zip(["A","B","C","D","E"], [30, 14, 10, 22, 18]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"


def _write_by_venue_sheet(ws, events: list[dict]):
    from openpyxl.styles import Font, PatternFill, Alignment
    from itertools import groupby

    ws.title = "By Venue"

    ws.merge_cells("A1:E1")
    ws["A1"] = "Events by Venue"
    ws["A1"].font = Font(bold=True, color=WHITE, name="Arial", size=12)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 26

    sorted_events = sorted(events, key=lambda e: (e.get("venue") or "zzz", e.get("date") or ""))
    row = 3
    current_venue = None

    for ev in sorted_events:
        venue = ev.get("venue") or "Unknown Venue"
        if venue != current_venue:
            current_venue = venue
            ws[f"A{row}"] = venue.upper()
            ws[f"A{row}"].font = Font(bold=True, color=WHITE, name="Arial", size=10)
            ws[f"A{row}"].fill = PatternFill("solid", fgColor=TEAL)
            ws.merge_cells(f"A{row}:E{row}")
            ws.row_dimensions[row].height = 20
            row += 1

        try:
            dt = datetime.strptime(ev.get("date") or "", "%Y-%m-%d")
            date_str = dt.strftime("%a %b %d")
        except Exception:
            date_str = ev.get("date") or ""

        ws[f"A{row}"] = ev.get("performer") or ""
        ws[f"B{row}"] = date_str
        ws[f"C{row}"] = ev.get("time_start") or ""
        ws[f"D{row}"] = ev.get("stage") or ""
        ws[f"E{row}"] = ev.get("name") or ""
        ws.row_dimensions[row].height = 16
        row += 1

    for col, w in zip(["A","B","C","D","E"], [24, 14, 8, 18, 32]):
        ws.column_dimensions[col].width = w


def generate_report(
    all_events: list[dict],
    changes: dict,
    run_id: str,
) -> Path:
    """
    Build and save the Excel workbook. Returns the output path.
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        raise RuntimeError("openpyxl not installed. Run: pip install openpyxl")

    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    out_path = EXPORTS_DIR / REPORT_NAME

    wb = Workbook()

    today = date.today().isoformat()
    upcoming = [e for e in all_events if (e.get("date") or "") >= today]

    # Sheet 1: All Events
    ws_all = wb.active
    _write_events_sheet(ws_all, all_events, "All Events")

    # Sheet 2: Upcoming
    ws_up = wb.create_sheet()
    _write_events_sheet(ws_up, upcoming, "Upcoming")

    # Sheet 3: Changes
    ws_changes = wb.create_sheet()
    _write_changes_sheet(ws_changes, changes)

    # Sheet 4: By Venue
    ws_venue = wb.create_sheet()
    _write_by_venue_sheet(ws_venue, all_events)

    wb.save(str(out_path))
    logger.info("Excel report saved to %s", out_path)
    return out_path
